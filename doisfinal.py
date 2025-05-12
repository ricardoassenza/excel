import pandas as pd
from openpyxl import load_workbook


def transferencia(caminho_origem, caminho_destino):
    aba_origem = 'base_filtrada'
    aba_destino = 'Base'

    dados_filtrados = pd.read_excel(caminho_origem, sheet_name=aba_origem)

    book = load_workbook(caminho_destino)

    if aba_destino not in book.sheetnames:
        book.create_sheet(aba_destino)

# Acessa a aba
    aba = book[aba_destino]

# Descobre a próxima linha vazia
    for i, row in enumerate(aba.iter_rows(values_only=True), 1):
        if any(cell not in (None, "") for cell in row):
            ultima_linha = i

# Ajusta para colar na linha seguinte
    linha_inicio = ultima_linha + 1

# Cola os dados, sem cabeçalho
    for r_idx, row in enumerate(dados_filtrados.values, start=linha_inicio):
        for c_idx, value in enumerate(row, start=1):
         aba.cell(row=r_idx, column=c_idx, value=value)

# Salva o arquivo
    book.save(caminho_destino)

    print(f'Dados serão colados a partir da linha: {linha_inicio}')
         