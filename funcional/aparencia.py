from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


def pintura(caminho, aba, cor_fundo, cor_letra):
    # Carrega o workbook
    wb = load_workbook(caminho)

    # Seleciona a aba desejada
    aba = wb[aba]

    # Define o preenchimento (fundo do cabeçalho)
    preenchimento = PatternFill(start_color=cor_fundo, end_color=cor_fundo, fill_type="solid")  # Azul claro

    # Define a fonte (cor do texto)
    fonte = Font(color=cor_letra, bold=True)  # Preto e em negrito

    # Aplica em cada célula da primeira linha (cabeçalho)
    for celula in aba[1]:
        celula.fill = preenchimento
        celula.font = fonte

    # Salva o arquivo
    wb.save(caminho)
    print('pintada')

