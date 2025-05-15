from openpyxl import load_workbook

def ordem_decrescente(caminho, aba):
    wb = load_workbook(caminho)
    ws = wb["Tabela"]

    header = [cell.value for cell in ws[1]]

    coluna_nome = None
    coluna_total = None

    for row in ws.iter_rows(min_row=1, max_row=10):
        for idx, cell in enumerate(row):
            valor = str(cell.value).lower() if cell.value else ""
            if "rótulos de linha" in valor:
                coluna_nome = idx
            if "total geral" in valor:
                coluna_total = idx
        if coluna_nome is not None and coluna_total is not None:
            break

    if coluna_nome is None or coluna_total is None:
        raise ValueError("Não foi possível encontrar as colunas 'Rótulos de Linha' e 'Total Geral'")


    dados = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        nome = row[coluna_nome]
        total = row[coluna_total]
        if nome and isinstance(total, (int, float)):
            dados.append((nome, total))


    dados_ordenados = sorted(dados, key=lambda x: x[1], reverse=False)


    if aba in wb.sheetnames:
        del wb[aba]
    ws_out = wb.create_sheet(aba)


    # Cabeçalho com 4 colunas
    ws_out.append(["Fornecedor", "Acum. 2025", "Frequência", "Finalidade"])

    # Adiciona os dados com colunas vazias para frequência e finalidade
    for item in dados_ordenados:
        ws_out.append([item[0], item[1], "", ""])

    def exclui_dados(item, coluna, aba):
        ws = wb[aba]
        for row in range(ws.max_row, 1, -1):
            status = ws.cell(row=row, column=coluna).value
            if status and item in status:
                ws.delete_rows(row)

    exclui_dados('BETZ', 1, aba)
    exclui_dados('SIEMENS', 1, aba)
    exclui_dados('Total Geral', 1, aba)
    exclui_dados('VENTOS PARAZINHENSES', 1, aba)
    exclui_dados('BRADESCO', 1, aba)
    exclui_dados('ITAU', 1, aba)
    exclui_dados('SERVENG CIVILSAN', 1, aba)
         

    wb.save(caminho)
    print('ordem decrescente finalizada')


