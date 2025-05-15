from openpyxl import load_workbook

def formatacao(caminho, aba, colum):
    # Abrir o arquivo Excel
    wb = load_workbook(caminho)
    ws = wb[aba]  # ou wb["NomeDaAba"]

    # Escolher a coluna, por exemplo, coluna B (índice 2)
    coluna = colum

    formato_contabil = 'R$ #,##0.00;-R$ #,##0.00'

    # Aplicar a formatação de contábil nas células da coluna
    for row in range(2, ws.max_row + 1):  # pula o cabeçalho (linha 1)
        cell = ws[f"{coluna}{row}"]
        if isinstance(cell.value, (int, float)):  # aplica só em números
            cell.number_format = formato_contabil

    # Salvar o arquivo
    wb.save(caminho)
    print('Formatado para contabil')