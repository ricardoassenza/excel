from openpyxl import load_workbook

caminho = r"C:\Users\52414463899\Documents\Copia\fornecedores2025_copia.xlsx"

wb = load_workbook(caminho,)
nome_aba = 'Base'
ws = wb[nome_aba]

for row in range(ws.max_row, 1, -1):
    status = ws.cell(row=row, column=25).value
    if status and 'REEMBOLSO' in status.upper():
        ws.delete_rows(row)

wb.save(caminho)        
print('ok')


    