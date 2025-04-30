from openpyxl import load_workbook

caminho = r"C:\Users\52414463899\Documents\Copia\controle-2025.xlsx"
nome_aba = 'Base'
coluna_filtro = 'Historico'
palavra_remover = 'REEMBOLSO'

# Abre o arquivo
wb = load_workbook(caminho)
ws = wb[nome_aba]

# Encontra o índice da coluna com nome 'Historico'
header = [cell.value for cell in ws[1]]
col_index = header.index(coluna_filtro) + 1  # openpyxl começa do 1

# Coleta as linhas que NÃO têm 'REEMBOLSO'
linhas_para_manter = []
for row in ws.iter_rows(min_row=2, values_only=False):  # min_row=2 para ignorar o cabeçalho
    valor_celula = row[col_index - 1].value
    if not (isinstance(valor_celula, str) and 'REEMBOLSO' in valor_celula.upper()):
        linhas_para_manter.append([cell.value for cell in row])

# Apaga todas as linhas abaixo do cabeçalho
ws.delete_rows(2, ws.max_row)

# Reinsere apenas as linhas válidas
for linha in linhas_para_manter:
    ws.append(linha)

# Salva a planilha
wb.save(caminho)
print('Linhas com "REEMBOLSO" removidas, fórmulas preservadas.')