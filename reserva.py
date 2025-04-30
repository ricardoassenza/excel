from openpyxl import load_workbook

caminho = r"C:\Users\52414463899\Documents\Copia\Controle _Fornecedores_2025_copia.xlsx"

wb = load_workbook(caminho)

def exclui_dados(item, coluna, aba):
    ws = wb[aba]
    for row in range(ws.max_row, 1, -1):
        status = ws.cell(row=row, column=coluna).value
        if status and item in status.upper():
            ws.delete_rows(row)

exclui_dados('REEMBOLSO', 25, 'Base')
exclui_dados('COMPANHIA', 26, 'Base')        
exclui_dados('Protheus', 29, 'Base')        

wb.save(caminho)        
print('ok')        

