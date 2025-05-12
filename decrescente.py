from openpyxl import load_workbook

def ordem_decrescente(caminho):
    wb = load_workbook(caminho)
    ws = wb["Tabela"]

    header = [cell.value for cell in ws[1]]

    coluna_nome = None
    coluna_total = None

    for row in ws.iter_rows(min_row=1, max_row=10):
        for idx, cell in enumerate(row):
            valor = str(cell.value).lower() if cell.value else ""
            if "rótulos de linha" in valor:
                coluna_nome = idx
            if "total geral" in valor:
                coluna_total = idx
        if coluna_nome is not None and coluna_total is not None:
            break

    if coluna_nome is None or coluna_total is None:
        raise ValueError("Não foi possível encontrar as colunas 'Rótulos de Linha' e 'Total Geral'")

    
    dados = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        nome = row[coluna_nome]
        total = row[coluna_total]
        if nome and isinstance(total, (int, float)):
            dados.append((nome, total))

    
    dados_ordenados = sorted(dados, key=lambda x: x[1], reverse=True)

    
    if "Ordenado" in wb.sheetnames:
        del wb["Ordenado"]
    ws_out = wb.create_sheet("Ordenado")
    ws_out.append(["fornecedores", "total"])

    for item in dados_ordenados:
        ws_out.append(item)

    def exclui_dados(item, coluna, aba):
        ws = wb[aba]
        for row in range(ws.max_row, 1, -1):
            status = ws.cell(row=row, column=coluna).value
            if status and item in status.upper():
                ws.delete_rows(row)

    exclui_dados('BETZ', 1, 'Ordenado')
    exclui_dados('SIEMENS', 1, 'Ordenado')     

    wb.save(caminho)
    print('ok')


